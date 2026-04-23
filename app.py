import streamlit as st
import openpyxl
from io import BytesIO

# 網頁配置
st.set_page_config(page_title="KPI 績效評分網頁系統", layout="wide")

# ==========================================
# 定義 Excel 座標 (根據您的附件精確設定)
# ==========================================
NAME_ROW = 2            # 員工姓名在第 2 列
WEIGHT_COL = 3          # 分數佔比在第 3 欄 (C)
START_EMP_COL = 4       # 第一個員工 (熊婷華) 在第 4 欄 (D)
# 指標列號定義
ROWS_1_TO_4 = [3, 4, 5, 6, 7, 8, 9, 10, 11]  # 評量指標 1~4
ROWS_5_TO_6 = [13, 14, 15, 16, 17, 18, 19]   # 加分項與扣分項 5~6
ROW_TEXT = 21                                # 加扣分事蹟文字列

def safe_float(val):
    try:
        if val is None or str(val).strip() == "": return 0.0
        return float(val)
    except (ValueError, TypeError):
        return 0.0

def format_to_percentage_string(val):
    try:
        if val is None or str(val).strip() == "": return "0%"
        f_val = float(val)
        return f"{f_val * 100:g}%"
    except ValueError:
        return str(val).strip()

if "all_scores" not in st.session_state:
    st.session_state.all_scores = {}

st.sidebar.title("📊 系統模式")
mode = st.sidebar.radio("切換功能：", ["模式 A：上傳 Excel 合併", "模式 B：網頁直接評分"])

# ---------------------------------------------------------
# 模式 A：上傳兩份 Excel 合併
# ---------------------------------------------------------
if mode == "模式 A：上傳 Excel 合併":
    st.title("📂 模式 A：自動合併兩位組長的 Excel")
    col1, col2 = st.columns(2)
    with col1: f1 = st.file_uploader("上傳【組長 A】檔案", type=["xlsx"], key="a1")
    with col2: f2 = st.file_uploader("上傳【組長 B】檔案", type=["xlsx"], key="a2")

    if f1 and f2:
        if st.button("執行合併計算"):
            try:
                # 讀取純數值來做計算，避免被公式干擾
                wb1_data = openpyxl.load_workbook(f1, data_only=True)
                wb2_data = openpyxl.load_workbook(f2, data_only=True)
                ws1_data = wb1_data.active
                ws2_data = wb2_data.active
                
                # 讀取模板用來覆寫並輸出 (保留格式)
                wb_out = openpyxl.load_workbook(f1) 
                ws_out = wb_out.active
                
                emp_cols = [c for c in range(START_EMP_COL, ws1_data.max_column + 1) if ws1_data.cell(row=NAME_ROW, column=c).value]
                
                for c in emp_cols:
                    # 【嚴格判定】：只加總 1~6 項的分數絕對值，> 0.0001 才算有評分
                    sum1 = sum(abs(safe_float(ws1_data.cell(row=r, column=c).value)) for r in ROWS_1_TO_4 + ROWS_5_TO_6)
                    sum2 = sum(abs(safe_float(ws2_data.cell(row=r, column=c).value)) for r in ROWS_1_TO_4 + ROWS_5_TO_6)
                    
                    # 1~4項 平均邏輯
                    for r in ROWS_1_TO_4:
                        v1 = safe_float(ws1_data.cell(row=r, column=c).value)
                        v2 = safe_float(ws2_data.cell(row=r, column=c).value)
                        
                        if sum1 > 1e-4 and sum2 > 1e-4:
                            ws_out.cell(row=r, column=c).value = (v1 + v2) / 2.0
                        elif sum2 > 1e-4:
                            ws_out.cell(row=r, column=c).value = v2 
                        elif sum1 > 1e-4:
                            ws_out.cell(row=r, column=c).value = v1
                        else:
                            ws_out.cell(row=r, column=c).value = 0.0
                    
                    # 5~6項 累加邏輯
                    for r in ROWS_5_TO_6:
                        v1 = safe_float(ws1_data.cell(row=r, column=c).value) if sum1 > 1e-4 else 0.0
                        v2 = safe_float(ws2_data.cell(row=r, column=c).value) if sum2 > 1e-4 else 0.0
                        ws_out.cell(row=r, column=c).value = v1 + v2
                    
                    # 文字合併 (不受是否打分限制)
                    t1 = str(ws1_data.cell(row=ROW_TEXT, column=c).value or "").strip()
                    t2 = str(ws2_data.cell(row=ROW_TEXT, column=c).value or "").strip()
                    if t1 and t2: ws_out.cell(row=ROW_TEXT, column=c).value = f"【組長A】:\n{t1}\n\n【組長B】:\n{t2}"
                    elif t2: ws_out.cell(row=ROW_TEXT, column=c).value = t2
                    elif t1: ws_out.cell(row=ROW_TEXT, column=c).value = t1

                out = BytesIO(); wb_out.save(out); out.seek(0)
                st.success("✅ 合併完成！未評分的組長已精準排除，不再拉低平均。")
                st.download_button("📥 下載結果", out, "KPI_合併結果.xlsx")
            except Exception as e:
                st.error(f"合併時發生錯誤，詳細錯誤：{e}")

# ---------------------------------------------------------
# 模式 B：線上直接填表評分
# ---------------------------------------------------------
else:
    st.title("📝 模式 B：組長線上評分系統")
    st.sidebar.subheader("第一步：設定範本")
    template = st.sidebar.file_uploader("上傳空白 Excel 範本", type=["xlsx"])

    if template:
        try:
            wb_temp = openpyxl.load_workbook(template, data_only=True)
            ws_temp = wb_temp.active
            
            employees = {}
            for c in range(START_EMP_COL, ws_temp.max_column + 1):
                name = ws_temp.cell(row=NAME_ROW, column=c).value
                if name: employees[str(name).strip()] = c
            
            item_info = {}
            for r in ROWS_1_TO_4 + ROWS_5_TO_6:
                cat = str(ws_temp.cell(row=r, column=1).value or "").strip()
                desc = str(ws_temp.cell(row=r, column=2).value or "").strip()
                weight_raw = ws_temp.cell(row=r, column=WEIGHT_COL).value
                label_text = f"{cat} {desc}".strip() if cat else desc
                item_info[r] = {
                    "label": label_text if label_text else f"未命名項目 (第 {r} 列)", 
                    "weight_str": format_to_percentage_string(weight_raw)
                }
                
            st.sidebar.success(f"已成功辨識：{len(employees)} 位員工")
            st.sidebar.write("員工名單：", ", ".join(employees.keys()))

            tab1, tab2 = st.tabs(["✍️ 組長評分填寫", "📥 結算與匯出"])
            
            with tab1:
                l_name = st.text_input("請輸入組長姓名：")
                e_name = st.selectbox("請選擇被評分員工：", ["--請選擇--"] + list(employees.keys()))
                
                if l_name and e_name != "--請選擇--":
                    st.markdown(f"### 🎯 正在為 **{e_name}** 評分")
                    st.info("💡 **貼心提示：** 請直接輸入百分比數字。例如想給 8.5% 分，請直接在格子輸入 **8.5** 即可。")
                    
                    with st.form(f"form_{l_name}_{e_name}"):
                        curr_scores = st.session_state.all_scores.get(l_name, {}).get(e_name, {})
                        new_data = {}
                        
                        st.subheader("1-4 項評量指標 (將採平均計算)")
                        for r in ROWS_1_TO_4:
                            info = item_info.get(r, {"label": f"項目載入失敗 (第 {r} 列)", "weight_str": "0%"})
                            st.write(f"📌 {info['label']}")
                            st.markdown(f"<p style='color:#007BFF; margin-bottom:5px;'>🎯 分數佔比上限：<b>{info['weight_str']}</b></p>", unsafe_allow_html=True)
                            display_val = float(curr_scores.get(r, 0.0)) * 100 
                            input_pct = st.number_input("輸入評分 (%)", value=display_val, format="%.2f", step=0.5, key=f"score_{r}_{e_name}")
                            new_data[r] = input_pct / 100.0
                        
                        st.subheader("5-6 項加減分 (將採累計計算)")
                        for r in ROWS_5_TO_6:
                            info = item_info.get(r, {"label": f"項目載入失敗 (第 {r} 列)", "weight_str": "0%"})
                            st.write(f"📌 {info['label']}")
                            st.markdown(f"<p style='color:#DC3545; margin-bottom:5px;'>🎯 加減分標準：<b>{info['weight_str']}</b></p>", unsafe_allow_html=True)
                            display_val = float(curr_scores.get(r, 0.0)) * 100
                            input_pct = st.number_input("輸入分數 (%)", value=display_val, format="%.2f", step=0.5, key=f"score_{r}_{e_name}")
                            new_data[r] = input_pct / 100.0
                            
                        st.subheader("加扣分事蹟說明")
                        new_data[ROW_TEXT] = st.text_area("請詳細填寫事蹟內容", value=curr_scores.get(ROW_TEXT, ""), height=100, key=f"txt_{e_name}")
                        
                        if st.form_submit_button("💾 儲存這份評分"):
                            if l_name not in st.session_state.all_scores: st.session_state.all_scores[l_name] = {}
                            st.session_state.all_scores[l_name][e_name] = new_data
                            st.success(f"✅ 已成功儲存 {l_name} 對 {e_name} 的評分！可以繼續選下一位。")

            with tab2:
                if not st.session_state.all_scores:
                    st.warning("⚠️ 目前系統內尚無任何組長的評分紀錄。")
                else:
                    st.write("📋 **目前收集到的組長評分進度：**")
                    for l in st.session_state.all_scores.keys():
                        st.write(f"- 👤 **{l}** 組長 (已評人數: {len(st.session_state.all_scores[l])} 人)")
                    
                    if st.button("🚀 開始計算 KPI 並產生 Excel", type="primary", use_container_width=True):
                        wb_out = openpyxl.load_workbook(template)
                        ws_out = wb_out.active
                        
                        for emp_name, col_idx in employees.items():
                            active_leaders = []
                            # 【嚴格判定】：只抓取 1~6 項分數加總 > 0.0001 的組長
                            for l in st.session_state.all_scores:
                                if emp_name in st.session_state.all_scores[l]:
                                    d = st.session_state.all_scores[l][emp_name]
                                    total_numeric_score = sum(abs(float(d.get(r, 0.0))) for r in ROWS_1_TO_4 + ROWS_5_TO_6)
                                    if total_numeric_score > 1e-4:
                                        active_leaders.append(l)
                            
                            num = len(active_leaders)
                            
                            # 1~4項 平均
                            for r in ROWS_1_TO_4:
                                if num > 0:
                                    total = sum(float(st.session_state.all_scores[l][emp_name].get(r, 0.0)) for l in active_leaders)
                                    ws_out.cell(row=r, column=col_idx).value = total / num
                                else:
                                    ws_out.cell(row=r, column=col_idx).value = 0.0
                                    
                            # 5~6項 累加
                            for r in ROWS_5_TO_6:
                                if num > 0:
                                    total = sum(float(st.session_state.all_scores[l][emp_name].get(r, 0.0)) for l in active_leaders)
                                    ws_out.cell(row=r, column=col_idx).value = total
                                else:
                                    ws_out.cell(row=r, column=col_idx).value = 0.0
                                    
                            # 【分離處理】：文字合併不受上述分數限制，即使沒打分只要有寫字都會抓進來
                            txts = []
                            for l in st.session_state.all_scores:
                                if emp_name in st.session_state.all_scores[l]:
                                    t = str(st.session_state.all_scores[l][emp_name].get(ROW_TEXT, "")).strip()
                                    if t: txts.append(f"【{l}】:\n{t}")
                            if txts:
                                ws_out.cell(row=ROW_TEXT, column=col_idx).value = "\n\n".join(txts)
                        
                        final_out = BytesIO(); wb_out.save(final_out); final_out.seek(0)
                        st.success("🎉 全員成績結算完畢！未評分者已精確排除。")
                        st.download_button("📥 下載最終 KPI 報表 Excel 檔", final_out, "KPI_最終網頁彙整結果.xlsx", use_container_width=True)
        except Exception as e:
            st.error(f"讀取範本時發生錯誤。錯誤資訊：{e}")
    else:
        st.info("👈 請先從左側邊欄上傳 Excel 範本以顯示評分表單。")
